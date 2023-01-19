# -*- coding: utf-8 -*-
#!/usr/bin/env python
import sys
reload(sys)
sys.setdefaultencoding('utf8')

import os
import importlib
import logging
# from xlwt import Workbook
from openpyxl import Workbook

from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText
from email.MIMEBase import MIMEBase
from email import encoders
import smtplib

from io import BytesIO

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lms.envs.aws")
os.environ.setdefault("lms.envs.aws,SERVICE_VARIANT", "lms")
os.environ.setdefault("PATH", "/edx/app/edxapp/venvs/edxapp/bin:/edx/app/edxapp/edx-platform/bin:/edx/app/edxapp/.rbenv/bin:/edx/app/edxapp/.rbenv/shims:/edx/app/edxapp/.gem/bin:/edx/app/edxapp/edx-platform/node_modules/.bin:/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin")

os.environ.setdefault("SERVICE_VARIANT", "lms")
os.chdir("/edx/app/edxapp/edx-platform")

startup = importlib.import_module("lms.startup")
startup.run()

log = logging.getLogger()

from opaque_keys.edx.keys import CourseKey

from student.models import CourseEnrollment, UserProfile

course_ids=[
    "course-v1:afpa+LaPatisserie+MOOCPatisserieAFPA_S1",
    "course-v1:afpa+LaPatisserie2+MOOCPatisserieAFPA_S2",
    "course-v1:afpa+MOOC_FLE_AFPA+FLE",
    "course-v1:afpa+Metsetvins+MOOCmetsetvinsAFPA_S3",
    "course-v1:afpa+Les101techniquesdebase+MOOCCUISINEAFPA",
    "course-v1:afpa+Les101techniquesdebase+MOOCCUISINEAFPA_S2",
    "course-v1:afpa+Les101techniquesdebase+MOOCCUISINEAFPA_S3",
    "course-v1:afpa+Les101techniquesreplay+2019",
    "course-v1:afpa+occitanie+2019_S1",
    "course-v1:afpa+MOOC_FLI+FLI_2019",
    "course-v1:afpa+La_Patisserie_replay_2020+2020",
    "course-v1:afpa+Mets_et_vins_replay_2020+2020",
    "course-v1:afpa+MOOC_FLI_replay_2020+2020",
    "course-v1:afpa+replay_2020+2020",
    "course-v1:afpa+mixite+mixite_2020",
    "course-v1:afpa+CPF+CPF_2020",
    "course-v1:afpa+inclusion_sociale+2020",
    "course-v1:afpa+TRE_2020+2020",
    "course-v1:afpa+MATU+2020",
    "course-v1:afpa+love_food+2020"
]

# Get all the user's information and set it in a list
users_list = []
for course in course_ids:
    course_key = CourseKey.from_string(course)
    enrollments = CourseEnrollment.objects.filter(course_id=course_key)

    for index, enrollment in enumerate(enrollments):
        user = enrollment.user
        username = str(user)
        user_email = str(user.email)
        registration_date = str(user.date_joined.strftime('%d/%m/%Y'))
        try:
            last_login = str(user.last_login.strftime('%d/%m/%Y'))
        except AttributeError:
            last_login = 'never_logged'
        users_list.append([username, user_email, registration_date, last_login] )
        # if index > 3:
        #     break

# write the csv file
headers = [u"Nom d'utilisateur", u"Email",u"Date d'inscription",u"Dernière connexion"]
wb = Workbook()
filepath = '/home/edxtma/afpa_users_info.csv'
sheet = wb.active
for index, header in enumerate(headers):
    sheet.cell(1, index+1, header)
for index, user_data in enumerate(users_list):
    for jindex, data in enumerate(user_data):
        sheet.cell(index+2, jindex+1, data)
output = BytesIO()
wb.save(output)
# wb.save(filepath)
attached_file = output.getvalue()


def send_mail(email, template):
    cc = ['dimitri.hoareau@themoocagency.com']
    html = template.encode('utf-8')
    part2 = MIMEText(html, 'html')
    fromaddr = '<no-reply@themoocagency.com>'
    addr = str(email)
    toaddrs = [addr] + cc
    msg = MIMEMultipart()
    msg['From'] = fromaddr
    # msg['To'] = ','.join(toaddrs)
    msg['To'] = addr
    msg['Subject'] = '[e-formation.artisanat] Nouvelle inscription sur]'

    attachment = attached_file
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment)
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= %s" % os.path.basename(filepath))
    msg.attach(part)

    server = smtplib.SMTP('mail3.themoocagency.com', 25)
    server.starttls()
    server.login('contact', 'waSwv6Eqer89')
    msg.attach(part2)
    text = msg.as_string()
    log.info('Email sent to '+ str(toaddrs))
    server.sendmail(fromaddr, toaddrs, text)
    server.quit()

recipient_mail = ['dimitri.hoareau@themoocagency.com']
mail_template = 'hello'
send_mail(recipient_mail, mail_template)