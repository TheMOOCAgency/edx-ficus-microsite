# -*- coding: utf-8 -*-
#!/usr/bin/env python
from __future__ import division

import sys
reload(sys)
sys.setdefaultencoding('utf8')

import os
import importlib
import csv
import time
import json
import string
from collections import OrderedDict
from io import BytesIO

import smtplib
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText
from email.MIMEBase import MIMEBase
from email import encoders

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lms.envs.aws")
os.environ.setdefault("lms.envs.aws,SERVICE_VARIANT", "lms")
os.environ.setdefault("PATH", "/edx/app/edxapp/venvs/edxapp/bin:/edx/app/edxapp/edx-platform/bin:/edx/app/edxapp/.rbenv/bin:/edx/app/edxapp/.rbenv/shims:/edx/app/edxapp/.gem/bin:/edx/app/edxapp/edx-platform/node_modules/.bin:/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin")

os.environ.setdefault("SERVICE_VARIANT", "lms")
os.chdir("/edx/app/edxapp/edx-platform")

startup = importlib.import_module("lms.startup")
startup.run()
#IMPORT TO
##RUN OUTSITE EDX
from django.core.management import execute_from_command_line
import django
##USE EDX FUNCTIONS
from opaque_keys.edx.keys import CourseKey
from lms.djangoapps.grades.new.course_grade import CourseGradeFactory, CourseGrade
from student.models import *
from courseware.courses import get_course_by_id
from tma_apps.best_grade.helpers import check_best_grade


from openpyxl import Workbook

from tma_apps.files_api.models import mongofiles
from tma_apps.models import TmaCourseEnrollment


import logging
log = logging.getLogger()



string_emails = sys.argv[1]
TO_EMAILS = string_emails.split(';')
try:
    course_ids = sys.argv[2].split(';')
except:
    pass


# Needed for specific TimeTracking
course_hashes_dict = {
    "course-v1:academie-digitale+FC_B20+2022": ["fa885f496101480187ad0d1b11222e19", "c0ed61dfe56448ad80c922b26a8c8b89", "4241479f8f3b4f3f9230011cd71e24de", "e97cee85f42745d8a6e5ab0bae396cdf", "b8fa1ebbe81b4bfc84c39ee17b581c1a"],
    "course-v1:academie-digitale+FC_20+2022" : ["49398e7edb6b4efeb90a3f57af2b6d0c", "6490e613f4b64419bfb8dd541a15183a", "f800fc13f85a4b8da4e5820eee0409d7", "4eb292036cf94b11a66676246ed3c488", "987d0de706364baaa740b299fa632195", "3c1973612b3544a5835bae7f7fdac102", "4987691195504d509c3e5a62d274b0a1"]
}
all_users_data = {}
course_names = []
course_names_html = []


for course_id in course_ids:

    course = get_course_by_id(CourseKey.from_string(course_id)) 
    course_names.append(course.display_name_with_default.encode('ascii', errors='xmlcharrefreplace'))
    course_names_html.append("<li>"+ course.display_name_with_default.encode('ascii', errors='xmlcharrefreplace')+"</li>")

    course_key = CourseKey.from_string(course_id)
    course = get_course_by_id(course_key)
    course_enrollments = CourseEnrollment.objects.filter(course_id=course_key)

    for i in range(len(course_enrollments)):
        user = course_enrollments[i].user
        tma_enrollment,is_exist=TmaCourseEnrollment.objects.get_or_create(course_enrollment_edx=course_enrollments[i])

        # if user.email.find("cyril.adolf") == -1 :
        if user.email.find("@yopmail") != -1 or user.email.find("@weup") != -1 or user.email.find("@themoocagency") != -1 :
            log.info('yopmail account')
            continue

        # Create a new user_data or use already defined one
        if str(user.id) in all_users_data :
            user_data = all_users_data[str(user.id)]
        else:
            user_data = {'general':{}}

        user_data[str(course_id)] = {}


        # Update object with user data without grades
        try:
            user_data['general']["first_name"] = json.loads(user.profile.custom_field)['first_name'].capitalize()
            user_data['general']["last_name"] = json.loads(user.profile.custom_field)['last_name'].capitalize()
        except:
            user_data['general']["first_name"] = "n/a"
            user_data['general']["last_name"] = "n/a"

        try:
            user_data['general']["registration_date"] = str(user.date_joined.strftime('%d/%m/%Y'))
        except:
            # should not occured
            user_data['general']["registration_date"] = "n/a"    

        try:
            user_data['general']["last_visit"] = str(user.last_login.strftime('%d/%m/%Y'))
        except:
            try:
                user_data['general']["last_visit"] = str(user.date_joined.strftime('%d/%m/%Y'))
            except:
                # should not occured
                user_data['general']["last_visit"] = "n/a"

        user_data['general']["id"] = user.id
        user_data['general']["email_address"] = user.email


        course_grade = CourseGradeFactory().create(user, course)
        user_grade = check_best_grade(user, course, force_best_grade=True)

        grade = []

        if 'section_breakdown' in user_grade.grade_value.keys():

            for section in user_grade.grade_value['section_breakdown']:
                if 'prominent' in section.keys():
                    grade_partial = section['percent'] * 100
                    module_name = section['category']
                    grade.append([module_name,grade_partial])


        timetracking_data = json.loads(tma_enrollment.detailed_time_tracking)

        course_hashes = course_hashes_dict[course_id]

        if timetracking_data != {}:
            timetracking = {}
            for chapter in course_hashes :

                if chapter in timetracking_data.keys():

                    chapter_time = timetracking_data[chapter]
                    timetracking[chapter] = round(chapter_time // 60)

            user_data[str(course_id)]['timetracking'] = timetracking
        else:
            user_data[str(course_id)]['timetracking'] = {}

        user_data[str(course_id)]['grade'] = grade


        all_users_data[str(user.id)] = user_data



# WORKBOOK
timestr = time.strftime("%Y_%m_%d")
wb = Workbook()
sheet = wb.active
sheet.title= course_names[0]
filename = '/home/edxtma/csv/{}_academie-digitale.xlsx'.format(timestr)

first = True

for idx, course_id in enumerate(course_ids):

    if not first:
        sheet = wb.create_sheet(course_names[idx])

    headers = ["Identifiant","Email","Prénom","Nom de famille","Date d'inscription","Dernière connexion"]

    # WRITE FILE
    for i, header in enumerate(headers):
        sheet.cell(1, i+1, header)

    first = False
    course_header = False
    j = 1
    for index, user in all_users_data.items():
        
        # Genral
        sheet.cell(j+1, 1, user["general"]["id"])
        sheet.cell(j+1, 2, user["general"]["email_address"])
        sheet.cell(j+1, 3, user["general"]["first_name"])
        sheet.cell(j+1, 4, user["general"]["last_name"])
        sheet.cell(j+1, 5, user["general"]["registration_date"])
        sheet.cell(j+1, 6, user["general"]["last_visit"])
        
        i = 7
        if course_id in user.keys():
            
            # TimeTracker
            success = True
            for section in course_hashes_dict[course_id]:

                if section in user[course_id]["timetracking"]:
                    sheet.cell(j+1, i, user[course_id]["timetracking"][section])
                else:
                    success = False
                i+=2

            # Grades
            i=7

            # Grade
            for chapter in user[course_id]["grade"]:
                sheet.cell(j+1, i+1, chapter[1])

                if chapter[1] == 0:
                    success = False

                # if not course_header :
                sheet.cell(1, i, 'Temps (min): '+str(chapter[0]))
                sheet.cell(1, i+1, 'Note (%): '+ str(chapter[0]))
                    # course_header = True
                i+=2

            sheet.cell(1, i, 'Attestation')
            sheet.cell(j+1, i, 'Oui') if success  else sheet.cell(j+1, i, 'Non') 

        j += 1


# SEND MAILS
output = BytesIO()
wb.save(output)
_files_values = output.getvalue()


course_names_html = ''.join(course_names_html)

html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en pièce jointe le rapport de note : "+ course_names_html +"<br/>Bonne r&eacute;ception<br>L'&eacute;quipe WeUp Learning<br/></p></body></html>"

part2 = MIMEText(html.encode('utf-8'), 'html', 'utf-8')

for i in range(len(TO_EMAILS)):
   fromaddr = "ne-pas-repondre@themoocagency.com"
   toaddr = str(TO_EMAILS[i])
   msg = MIMEMultipart()
   msg['From'] = fromaddr
   msg['To'] = toaddr
   msg['Subject'] = "Academie-Digitale - " + ' + '.join(course_names)
   attachment = _files_values
   part = MIMEBase('application', 'octet-stream')
   part.set_payload(attachment)
   encoders.encode_base64(part)
   part.add_header('Content-Disposition', "attachment; filename= %s" % os.path.basename(filename))
   msg.attach(part)
   server = smtplib.SMTP('mail3.themoocagency.com', 25)
   server.starttls()
   server.login('contact', 'waSwv6Eqer89')
   msg.attach(part2)
   text = msg.as_string()
   server.sendmail(fromaddr, toaddr, text)
   server.quit()
   log.info('Email sent to '+str(TO_EMAILS[i]))


 
# Test
# sudo -H -u edxapp /edx/bin/python.edxapp /edx/app/edxapp/edx-microsite/academie-digitale/utils/time_tracking_grade_report.py "cyril.adolf@weuplearning.com" "course-v1:academie-digitale+FC_20+2022;course-v1:academie-digitale+FC_B20+2022" 


