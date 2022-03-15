# -*- coding: utf-8 -*-
#!/usr/bin/env python
from __future__ import division

import sys
reload(sys)
sys.setdefaultencoding('utf8')

import os
import importlib
import csv
import time
import json
import string
from collections import OrderedDict
from io import BytesIO

import smtplib
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText
from email.MIMEBase import MIMEBase
from email import encoders

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lms.envs.aws")
os.environ.setdefault("lms.envs.aws,SERVICE_VARIANT", "lms")
os.environ.setdefault("PATH", "/edx/app/edxapp/venvs/edxapp/bin:/edx/app/edxapp/edx-platform/bin:/edx/app/edxapp/.rbenv/bin:/edx/app/edxapp/.rbenv/shims:/edx/app/edxapp/.gem/bin:/edx/app/edxapp/edx-platform/node_modules/.bin:/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin")

os.environ.setdefault("SERVICE_VARIANT", "lms")
os.chdir("/edx/app/edxapp/edx-platform")

startup = importlib.import_module("lms.startup")
startup.run()
#IMPORT TO
##RUN OUTSITE EDX
from django.core.management import execute_from_command_line
import django
##USE EDX FUNCTIONS
from opaque_keys.edx.keys import CourseKey
from lms.djangoapps.grades.new.course_grade import CourseGradeFactory, CourseGrade
from student.models import *
from courseware.courses import get_course_by_id
from tma_apps.best_grade.helpers import check_best_grade


from datetime import datetime
from openpyxl import Workbook

from tma_apps.files_api.models import mongofiles
from tma_apps.models import TmaCourseEnrollment


import logging
log = logging.getLogger()


string_emails = sys.argv[1]
TO_EMAILS = string_emails.split(';')
try:
    course_ids = sys.argv[2].split(';')
except:
    pass


all_users_data = {}


for course_id in course_ids:
    log.info(course_id)
    course_key = CourseKey.from_string(course_id)

    course = get_course_by_id(course_key)
    course_enrollments = CourseEnrollment.objects.filter(course_id=course_key)

    for i in range(len(course_enrollments)):
        user = course_enrollments[i].user
        tma_enrollment,is_exist=TmaCourseEnrollment.objects.get_or_create(course_enrollment_edx=course_enrollments[i])

        if user.email.find("@yopmail") != -1 or user.email.find("@weup") != -1 or user.email.find("@themoocagency") != -1 :
            log.info('yopmail account')
            continue

        # Create a new user_data or use already defined one
        if str(user.id) in all_users_data :
            user_data = all_users_data[str(user.id)]
        else:
            user_data = {'general':{}}

        user_data[str(course_id)] = {}


        # Update object with user data without grades
        try:
            user_data['general']["first_name"] = json.loads(user.profile.custom_field)['first_name'].capitalize()
            user_data['general']["last_name"] = json.loads(user.profile.custom_field)['last_name'].capitalize()
        except:
            user_data['general']["first_name"] = "n/a"
            user_data['general']["last_name"] = "n/a"

        try:
            user_data['general']["registration_date"] = str(datetime.strptime(user.date_joined.strftime('%d/%m/%Y'), '%d/%m/%Y') )
        except:
            # should not occured
            user_data['general']["registration_date"] = "n/a"    

        try:
            user_data['general']["last_visit"] = str(datetime.strptime(user.last_login.strftime('%d/%m/%Y'), '%d/%m/%Y'))
        except:
            try:
                user_data['general']["last_visit"] = str(datetime.strptime(user.date_joined.strftime('%d/%m/%Y'), '%d/%m/%Y'))
            except:
                # should not occured
                user_data['general']["last_visit"] = "n/a"

        user_data['general']["id"] = user.id
        user_data['general']["email_address"] = user.email

        # Finished course date
        # log.info("tma_enrollment.finished_course_date")
        # log.info(tma_enrollment.finished_course_date)

        course_grade = CourseGradeFactory().create(user, course)
        user_grade = check_best_grade(user, course, force_best_grade=True)

        # Only for one course 
        if course_id == 'course-v1:asffor+ASF_01+ASF_2020':
            grade = dict()
            for m in user_grade.grade_value['grade_breakdown'].keys():
                grade_partial = round(user_grade.grade_value['grade_breakdown'].get(m)['percent'] * 100, 0)
                module_name = user_grade.grade_value['grade_breakdown'].get(m)['category']
                #grade.append({module_name : grade_partial})
                grade[module_name] = grade_partial
                # log.info(tma_enrollment.detailed_time_tracking)
                # try:
                #     seconds = tma_enrollment.detailed_time_tracking
                #     user_data[str(course_id)]['time_tracking'] = tma_enrollment.detailed_time_tracking
                # except:
                #     user_data["time_tracking"] = int(0)

        else:
            grade = round(user_grade.percent *100 , 0)
            # try:
            #     seconds = tma_enrollment.global_time_tracking
            #     minute = seconds // 60
            #     user_data[str(course_id)]['time_tracking'] = int(minute)
            # except:
            #     user_data["time_tracking"] = int(0)


        user_data[str(course_id)]['grade'] = grade
        all_users_data[str(user.id)] = user_data


def Getletterfromindex(num):
    #produces a string from numbers so  1->A , 26->Z , 54->BB

    num2alphadict = dict(zip(range(1, 27), string.ascii_uppercase))
    outval = ""
    numloops = (num-1) //26
    if numloops > 0:
        outval = outval + Getletterfromindex(numloops)
    remainder = num % 26
    if remainder > 0:
        outval = outval + num2alphadict[remainder]
    else:
        outval = outval + "Z"
    return outval


headers = ["Identifiant","Email","Prénom", "Nom de famille","Date d'inscription","Dernière connexion",  "01_commercialisation", "02_eidas", '03_gouvernance', '04_acpr', '05_secret_medical', '06_cadre_reglementaire', '07_distribution', '08_fraude', '09_reclamations', '10_vad', '11_sc_M1','11_sc_M2','11_sc_M3','11_sc_M4','11_sc_M5','11_sc_M6','11_sc_M7','11_sc_M8','11_sc_M9','11_sc_M10','11_sc_M11','11_sc_M12','11_sc_M13','11_sc_M14','11_sc_M15','11_sc_M16','11_sc_M17','11_sc_M18','12_blanchiment' , 'Final grade' , 'Attestation']
headers_2 = ["Email","Prénom", "Nom de famille", "Temps (en heures)" , "01_commercialisation", "02_eidas", '03_gouvernance', '04_acpr', '05_secret_medical', '06_cadre_reglementaire', '07_distribution', '08_fraude', '09_reclamations', '10_vad', '11_sc_M1','11_sc_M2','11_sc_M3','11_sc_M4','11_sc_M5','11_sc_M6','11_sc_M7','11_sc_M8','11_sc_M9','11_sc_M10','11_sc_M11','11_sc_M12','11_sc_M13','11_sc_M14','11_sc_M15','11_sc_M16','11_sc_M17','11_sc_M18','12_blanchiment']

# WORKBOOK
timestr = time.strftime("%Y_%m_%d")
wb = Workbook()
sheet = wb.active
sheet.title= 'Grade Report'
sheet_2 = wb.create_sheet('Reporting en heure')
filename = '/home/edxtma/csv/{}_asffor.xlsx'.format(timestr)



# WRITE FILE
j = 1
i = 1
for i, header in enumerate(headers):
    sheet.cell(j, i+1, header)

for i, header in enumerate(headers_2):
    sheet_2.cell(j, i+1, header)


for index, user in all_users_data.items():
    # GENRAL
    sheet.cell(j+1, 1, user["general"]["id"])
    sheet.cell(j+1, 2, user["general"]["email_address"])
    sheet.cell(j+1, 3, user["general"]["first_name"])
    sheet.cell(j+1, 4, user["general"]["last_name"])
    try:
        sheet.cell(j+1, 5, user["general"]["registration_date"])
    except:
        sheet.cell(j+1, 5, "n.a.")
    try:
        sheet.cell(j+1, 6, user["general"]["last_visit"])
    except:
        sheet.cell(j+1, 6, "n.a.")
    
    # GRADES 
    
    # final grade
    final_grade = 0

    try: 
        sheet.cell(j+1, 7, user['course-v1:asffor+01+DDA_2022']['grade']) 
        sheet_2.cell(j+1, 5, 30) if user['course-v1:asffor+01+DDA_2022']['grade'] >= 80 else sheet_2.cell(j+1, 5, 0)
        final_grade += user['course-v1:asffor+01+DDA_2022']['grade']
    except: 
        sheet.cell(j+1, 7, 0)
        sheet_2.cell(j+1, 5, 0)
    
    try:
        sheet.cell(j+1, 8, user['course-v1:asffor+02+DDA_2022']['grade']) 
        sheet_2.cell(j+1, 6, 30) if user['course-v1:asffor+02+DDA_2022']['grade'] >= 80 else sheet_2.cell(j+1, 6, 0)
        final_grade += user['course-v1:asffor+02+DDA_2022']['grade']
    except:
        sheet.cell(j+1, 8, 0)
        sheet_2.cell(j+1, 6, 0)

    try:
        sheet.cell(j+1, 9, user['course-v1:asffor+03+DDA_2022']['grade']) 
        sheet_2.cell(j+1, 7, 30) if user['course-v1:asffor+03+DDA_2022']['grade'] >= 80 else sheet_2.cell(j+1, 7, 0)
        final_grade += user['course-v1:asffor+03+DDA_2022']['grade']
    except:
        sheet.cell(j+1, 9, 0)
        sheet_2.cell(j+1, 7, 0)

    try:
        sheet.cell(j+1, 10, user['course-v1:asffor+04+DDA_2022']['grade'])
        sheet_2.cell(j+1, 8, 30) if user['course-v1:asffor+04+DDA_2022']['grade'] >= 80 else sheet_2.cell(j+1, 8, 0)
        final_grade += user['course-v1:asffor+04+DDA_2022']['grade']
    except: 
        sheet.cell(j+1, 10, 0)
        sheet_2.cell(j+1, 8, 0)

    try:
        sheet.cell(j+1, 11, user['course-v1:asffor+05+DDA_2022']['grade'])
        sheet_2.cell(j+1, 9, 40) if user['course-v1:asffor+05+DDA_2022']['grade'] >= 80 else sheet_2.cell(j+1, 9, 0)
        final_grade += user['course-v1:asffor+05+DDA_2022']['grade']
    except:
        sheet.cell(j+1, 11, 0)
        sheet_2.cell(j+1, 9, 0)

    try:
        sheet.cell(j+1, 12, user['course-v1:asffor+06+DDA_2022']['grade']) 
        sheet_2.cell(j+1, 10, 25) if user['course-v1:asffor+06+DDA_2022']['grade'] >= 80 else sheet_2.cell(j+1, 10, 0)
        final_grade += user['course-v1:asffor+06+DDA_2022']['grade']
    except:
        sheet.cell(j+1, 12, 0)
        sheet_2.cell(j+1, 10, 0)

    try:
        sheet.cell(j+1, 13, user['course-v1:asffor+07+DDA_2022']['grade']) 
        sheet_2.cell(j+1, 11, 35) if user['course-v1:asffor+07+DDA_2022']['grade'] >= 80 else sheet_2.cell(j+1, 11, 0)
        final_grade += user['course-v1:asffor+07+DDA_2022']['grade']
    except:
        sheet.cell(j+1, 13, 0)
        sheet_2.cell(j+1, 11, 0)

    try:
        sheet.cell(j+1, 14, user['course-v1:asffor+08+DDA_2022']['grade'])
        sheet_2.cell(j+1, 12, 30) if user['course-v1:asffor+08+DDA_2022']['grade'] >= 80 else sheet_2.cell(j+1, 12, 0)
        final_grade += user['course-v1:asffor+08+DDA_2022']['grade']
    except:
        sheet.cell(j+1, 14, 0)
        sheet_2.cell(j+1, 12, 0)

    try:
        sheet.cell(j+1, 15, user['course-v1:asffor+09+DDA_2022']['grade']) 
        sheet_2.cell(j+1, 13, 20) if user['course-v1:asffor+09+DDA_2022']['grade'] >= 80 else sheet_2.cell(j+1, 13, 0)
        final_grade += user['course-v1:asffor+09+DDA_2022']['grade']
    except:
        sheet.cell(j+1, 15, 0)
        sheet_2.cell(j+1, 13, 0)

    try:
        sheet.cell(j+1, 16, user['course-v1:asffor+10+DDA_2022']['grade']) 
        sheet_2.cell(j+1, 14, 30) if user['course-v1:asffor+10+DDA_2022']['grade'] >= 80 else sheet_2.cell(j+1, 14, 0)
        final_grade += user['course-v1:asffor+10+DDA_2022']['grade']
    except:
        sheet.cell(j+1, 16, 0)
        sheet_2.cell(j+1, 14, 0)

    try:
        sheet.cell(j+1, 17, user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 1"]) 
        sheet_2.cell(j+1, 15, 30) if user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 1"] >= 60 else sheet_2.cell(j+1, 15, 0)
        final_grade += user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 1"]
    except:
        sheet.cell(j+1, 17, 0)
        sheet_2.cell(j+1, 15, 0)

    try:
        sheet.cell(j+1, 18, user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 2"])
        sheet_2.cell(j+1, 16, 30) if user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 2"] >= 60 else sheet_2.cell(j+1, 16, 0)
        final_grade += user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 2"]
    except:
        sheet.cell(j+1, 18, 0)
        sheet_2.cell(j+1, 16, 0)


    try:
        sheet.cell(j+1, 19, user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 3"]) 
        sheet_2.cell(j+1, 17, 30) if user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 3"] >= 60 else sheet_2.cell(j+1, 17, 0)
        final_grade += user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 3"]
    except:
        sheet.cell(j+1, 19, 0)
        sheet_2.cell(j+1, 17, 0)

    try:
        sheet.cell(j+1, 20, user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 4"])
        sheet_2.cell(j+1, 18, 30) if user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 4"] >= 60 else sheet_2.cell(j+1, 18, 0)
        final_grade += user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 4"]
    except:
        sheet.cell(j+1, 20, 0)
        sheet_2.cell(j+1, 18, 0)

    try:
        sheet.cell(j+1, 21, user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 5"]) 
        sheet_2.cell(j+1, 19, 30) if user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 5"] >= 60 else sheet_2.cell(j+1, 19, 0)
        final_grade += user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 5"]
    except:
        sheet.cell(j+1, 21, 0)
        sheet_2.cell(j+1, 19, 0)

    try:
        sheet.cell(j+1, 22, user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 6"]) 
        sheet_2.cell(j+1, 20, 30) if user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 6"] >= 60 else sheet_2.cell(j+1, 20, 0)
        final_grade += user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 6"]
    except:
        sheet.cell(j+1, 22, 0)
        sheet_2.cell(j+1, 20, 0)

    try:
        sheet.cell(j+1, 23, user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 7"])
        sheet_2.cell(j+1, 21, 30) if user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 7"] >= 60 else sheet_2.cell(j+1, 21, 0)
        final_grade += user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 7"]
    except:
        sheet.cell(j+1, 23, 0)
        sheet_2.cell(j+1, 21, 0)

    try:
        sheet.cell(j+1, 24, user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 8"])
        sheet_2.cell(j+1, 22, 30) if user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 8"] >= 60 else sheet_2.cell(j+1, 22, 0)
        final_grade += user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 8"]
    except:
        sheet.cell(j+1, 24, 0)
        sheet_2.cell(j+1, 22, 0)

    try:
        sheet.cell(j+1, 25, user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 9"])
        sheet_2.cell(j+1, 23, 30) if user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 9"] >= 60 else sheet_2.cell(j+1, 23, 0)
        final_grade += user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 9"]
    except:
        sheet.cell(j+1, 25, 0)
        sheet_2.cell(j+1, 23, 0)

    try:
        sheet.cell(j+1, 26, user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 10"]) 
        sheet_2.cell(j+1, 24, 30) if user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 10"] >= 60 else sheet_2.cell(j+1, 24, 0)
        final_grade += user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 10"]
    except:
        sheet.cell(j+1, 26, 0)
        sheet_2.cell(j+1, 24, 0)

    try:
        sheet.cell(j+1, 27, user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 11"])
        sheet_2.cell(j+1, 25, 30) if user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 11"] >= 60 else sheet_2.cell(j+1, 25, 0)
        final_grade += user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 11"]
    except:
        sheet.cell(j+1, 27, 0)
        sheet_2.cell(j+1, 25, 0)

    try:
        sheet.cell(j+1, 28, user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 12"])
        sheet_2.cell(j+1, 26, 30) if user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 12"] >= 60 else sheet_2.cell(j+1, 26, 0)
        final_grade += user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 12"]
    except:
        sheet.cell(j+1, 28, 0)
        sheet_2.cell(j+1, 26, 0)

    try:
        sheet.cell(j+1, 29, user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 13"])
        sheet_2.cell(j+1, 27, 30) if user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 13"] >= 60 else sheet_2.cell(j+1, 27, 0)
        final_grade += user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 13"]
    except:
        sheet.cell(j+1, 29, 0)
        sheet_2.cell(j+1, 27, 0)

    try:
        sheet.cell(j+1, 30, user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 14"])
        sheet_2.cell(j+1, 28, 30) if user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 14"] >= 60 else sheet_2.cell(j+1, 28, 0)
        final_grade += user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 14"]
    except:
        sheet.cell(j+1, 30, 0)
        sheet_2.cell(j+1, 28, 0)

    try:
        sheet.cell(j+1, 31, user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 15"])
        sheet_2.cell(j+1, 29, 30) if user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 15"] >= 60 else sheet_2.cell(j+1, 29, 0)
        final_grade += user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 15"]
    except:
        sheet.cell(j+1, 31, 0)
        sheet_2.cell(j+1, 29, 0)

    try:
        sheet.cell(j+1, 32, user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 16"])
        sheet_2.cell(j+1, 30, 30) if user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 16"] >= 60 else sheet_2.cell(j+1, 30, 0)
        final_grade += user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 16"]
    except:
        sheet.cell(j+1, 32, 0)
        sheet_2.cell(j+1, 30, 0)

    try:
        sheet.cell(j+1, 33, user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 17"])
        sheet_2.cell(j+1, 31, 30) if user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 17"] >= 60 else sheet_2.cell(j+1, 31, 0)
        final_grade += user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 17"]
    except:
        sheet.cell(j+1, 33, 0)
        sheet_2.cell(j+1, 31, 0)

    try:
        sheet.cell(j+1, 34, user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 18"]) 
        sheet_2.cell(j+1, 32, 30) if user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 18"] >= 60 else sheet_2.cell(j+1, 32, 0)
        final_grade += user['course-v1:asffor+ASF_01+ASF_2020']['grade']["Module 18"]
    except:
        sheet.cell(j+1, 34, 0)
        sheet_2.cell(j+1, 32, 0)

    try:
        sheet.cell(j+1, 35, user['course-v1:asffor+12+DDA_2022']['grade']) if user['course-v1:asffor+12+DDA_2022'] else sheet.cell(j+1, 35, 0)
        sheet_2.cell(j+1, 33, 60) if user['course-v1:asffor+10+DDA_2022']['grade'] >= 80 else sheet_2.cell(j+1, 33, 0)
        final_grade += user['course-v1:asffor+12+DDA_2022']['grade']
    except:
        sheet.cell(j+1, 35, 0)
        sheet_2.cell(j+1, 33, 0)

    sheet.cell(j+1, 36, final_grade//29 )
    sheet.cell(j+1, 37, 'oui') if final_grade >= (80*29) else sheet.cell(j+1, 37, 'non') 


    # PAGE 2
    sheet_2.cell(j+1, 1, user["general"]["email_address"])
    sheet_2.cell(j+1, 2, user["general"]["first_name"])
    sheet_2.cell(j+1, 3, user["general"]["last_name"])

    total_time = ""
    i = 5
    # course_ids - 1 + 18 (modules) 
    for grade in range(29):

        total_time += Getletterfromindex(i) + str(j+1)
        total_time += " + "
        i += 1
    total_time = total_time[0:-2]
    sheet_2.cell(j+1, 4, ("=SUM("+total_time+") / 60" ))

    j += 1


# SEND MAILS
output = BytesIO()
wb.save(output)
_files_values = output.getvalue()

course_names = []
course_names_html = []
for course_id in course_ids: 
    course = get_course_by_id(CourseKey.from_string(course_id)) 
    course_names.append(course.display_name_with_default.encode('ascii', errors='xmlcharrefreplace'))
    course_names_html.append("<li>"+ course.display_name_with_default.encode('ascii', errors='xmlcharrefreplace')+"</li>")

course_names_html = ''.join(course_names_html)

html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en pièce jointe le rapport de note : "+ course_names_html +"<br/>Bonne r&eacute;ception<br>L'&eacute;quipe WeUp Learning<br/></p></body></html>"

part2 = MIMEText(html.encode('utf-8'), 'html', 'utf-8')

for i in range(len(TO_EMAILS)):
   fromaddr = "ne-pas-repondre@themoocagency.com"
   toaddr = str(TO_EMAILS[i])
   msg = MIMEMultipart()
   msg['From'] = fromaddr
   msg['To'] = toaddr
   msg['Subject'] = "ASFFOR - " + ' + '.join(course_names)
   attachment = _files_values
   part = MIMEBase('application', 'octet-stream')
   part.set_payload(attachment)
   encoders.encode_base64(part)
   part.add_header('Content-Disposition', "attachment; filename= %s" % os.path.basename(filename))
   msg.attach(part)
   server = smtplib.SMTP('mail3.themoocagency.com', 25)
   server.starttls()
   server.login('contact', 'waSwv6Eqer89')
   msg.attach(part2)
   text = msg.as_string()
   server.sendmail(fromaddr, toaddr, text)
   server.quit()
   log.info('Email sent to '+str(TO_EMAILS[i]))


 
# Feb 2022
# command pprod1 
# sudo -H -u edxapp /edx/bin/python.edxapp /edx/app/edxapp/edx-microsite/asffor/utils/asffor_grade_report.py "cyril.adolf@weuplearning.com" "course-v1:asffor+ASF_01+ASF_2019;course-v1:asffor+ASF_02+ASF_2019" 

# command prod 

# sudo -H -u edxapp /edx/bin/python.edxapp /edx/app/edxapp/edx-microsite/asffor/utils/asffor_grade_report.py "cyril.adolf@weuplearning.com" "course-v1:asffor+01+DDA_2022;course-v1:asffor+ASF_01+ASF_2020" 

# sudo -H -u edxapp /edx/bin/python.edxapp /edx/app/edxapp/edx-microsite/asffor/utils/asffor_grade_report.py "cyril.adolf@weuplearning.com;manal.touati@weuplearning.com" "course-v1:asffor+05+DDA_2022;course-v1:asffor+ASF_01+ASF_2020" 

# CRONTAB: 
# 0 8 * * MON sudo -H -u edxapp /edx/bin/python.edxapp /edx/app/edxapp/edx-microsite/asffor/utils/asffor_grade_report.py "cyril.adolf@weuplearning.com;manal.touati@weuplearning.com;j.bontemps@asf-france.com;jsoulie@asf-france.com;a.matsakis@asf-france.com;m.malivert@asf-france.com" "course-v1:asffor+01+DDA_2022;course-v1:asffor+02+DDA_2022;course-v1:asffor+03+DDA_2022;course-v1:asffor+04+DDA_2022;course-v1:asffor+05+DDA_2022;course-v1:asffor+06+DDA_2022;course-v1:asffor+07+DDA_2022;course-v1:asffor+08+DDA_2022;course-v1:asffor+09+DDA_2022;course-v1:asffor+10+DDA_2022;course-v1:asffor+12+DDA_2022;course-v1:asffor+ASF_01+ASF_2020" 
