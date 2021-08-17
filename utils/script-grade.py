# -*- coding: utf-8 -*-
#!/usr/bin/env python
from __future__ import division

import sys
reload(sys)
sys.setdefaultencoding('utf8')

#IMPORT FOR SCRIPT TO
##RUN
##WRITE EXCEL FILE
##SEND EMAIL
import os
import importlib
import csv
import time
import os
import json
import logging
import string
from collections import OrderedDict



from io import BytesIO

import smtplib
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText
from email.MIMEBase import MIMEBase
from email import encoders

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lms.envs.aws")
os.environ.setdefault("lms.envs.aws,SERVICE_VARIANT", "lms")
os.environ.setdefault("PATH", "/edx/app/edxapp/venvs/edxapp/bin:/edx/app/edxapp/edx-platform/bin:/edx/app/edxapp/.rbenv/bin:/edx/app/edxapp/.rbenv/shims:/edx/app/edxapp/.gem/bin:/edx/app/edxapp/edx-platform/node_modules/.bin:/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin")

os.environ.setdefault("SERVICE_VARIANT", "lms")
os.chdir("/edx/app/edxapp/edx-platform")

startup = importlib.import_module("lms.startup")
startup.run()
#IMPORT TO
##RUN OUTSITE EDX
from django.core.management import execute_from_command_line
import django
##USE EDX FUNCTIONS
from opaque_keys.edx.keys import CourseKey
from courseware.access import has_access
from lms.djangoapps.ccx.utils import prep_course_for_grading
from lms.djangoapps.courseware import courses
#from lms.djangoapps.courseware.exceptions import CourseAccessRedirect
from lms.djangoapps.grades.api.serializers import GradingPolicySerializer
from lms.djangoapps.grades.new.course_grade import CourseGradeFactory, CourseGrade
from lms.djangoapps.tma_stat_dashboard.grade_reports import grade_reports
from openedx.core.lib.api.view_utils import DeveloperErrorViewMixin, view_auth_classes
from openedx.core.djangoapps.course_groups.models import CohortMembership, CourseUserGroup
from openedx.core.djangoapps.content.course_overviews.models import CourseOverview
from student.roles import CourseStaffRole
from student.models import *
from courseware.courses import get_course_by_id
from courseware.courses import get_course
from tma_ensure_form.models import ensure_form_models
from tma_apps.best_grade.helpers import check_best_grade
from lms.djangoapps.grades.context import grading_context_for_course
from courseware.user_state_client import DjangoXBlockUserStateClient



from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from microsite_configuration.models import (
    MicrositeOrganizationMapping,
    Microsite
)
from tma_apps.files_api.models import mongofiles
from tma_apps.models import TmaCourseEnrollment

log = logging.getLogger()

string_emails = sys.argv[1]
TO_EMAILS = string_emails.split(';')
try:
    course_ids = sys.argv[2].split(';')
except:
    pass

# WORKBOOK
timestr = time.strftime("%Y_%m_%d")
timesfr = time.strftime("%d.%m.%Y")
timesfr = str(timesfr)
wb = Workbook()
# wb = Workbook(encoding='utf-8')
sheet = wb.active
sheet.title= 'Rapport'
filename = '/home/edxtma/csv/{}_ntp_Engie.xls'.format(timestr)

# Format date
# date_style = Style(number_format='DD/MM/YYYY')
# style1.num_format_str = 'DD/MM/YYYY'

# Course_ids must be from same platform : getting microsite via first course_id in list
course_key = CourseKey.from_string(course_ids[0])
course = get_course_by_id(course_key)
# Get microsite
org = course.org
query = "SELECT a.id,a.organization,b.key FROM microsite_configuration_micrositeorganizationmapping a,microsite_configuration_microsite b WHERE a.microsite_id = b.id"
microsite_list = MicrositeOrganizationMapping.objects.raw(query)
microsite_name = None
for row in microsite_list:
    if row.organization == org:
        microsite_name = row.key
domain_prefix = None
register_form = None
certificate_form = None
microsite = Microsite.objects.get(key=microsite_name)
microsite_value = microsite.values
i=0
for val in microsite_value:
    if val == 'domain_prefix':
        domain_prefix = microsite_value.values()[i]
    if val == 'FORM_EXTRA':
        register_form = microsite_value.values()[i]
    if val == 'CERTIFICATE_FORM_EXTRA':
        certificate_form = microsite_value.values()[i]
    i=i+1
# Mongo certificate_form models
_mongo = ensure_form_models()
db = 'ensure_form'
collection = 'certificate_form'
_mongo.connect(db=db,collection=collection)
_mongo.microsite = domain_prefix


all_users_data = {}


for course_id in course_ids:
    course_key = CourseKey.from_string(course_id)
    course_overview = CourseOverview.get_from_id(course_key)
    course = get_course_by_id(course_key)
    course_enrollments = CourseEnrollment.objects.filter(course_id=course_key)
    
    # add course headers
    user_repports_summary = course_enrollments[0]
    user_summary=user_repports_summary.user



    for i in range(len(course_enrollments)):
        user = course_enrollments[i].user
        tma_enrollment,is_exist=TmaCourseEnrollment.objects.get_or_create(course_enrollment_edx=course_enrollments[i])

        # Create a new user_data
        user_data = {}        

        # number of connections        
        # connections = 1
        # try:
        #     user_ip = UserIpTracking.objects.filter(user=user,microsite=domain_prefix)
        #     for ip in user_ip:
        #         connections = connections + 1
        # except:
        #     pass
        # user_data["connections"] = connections

        # Update object with user data without grades
        try:
            user_data["first_name"] = json.loads(user.profile.custom_field)['first_name'].capitalize()
            user_data["last_name"] = json.loads(user.profile.custom_field)['last_name'].capitalize()
        except:
            user_data["first_name"] = "n/a"
            user_data["last_name"] = "n/a"

        try:
            user_data["registration_date"] = str(datetime.strptime(user.date_joined.strftime('%d/%m/%Y'), '%d/%m/%Y') )
        except:
            # should not occured
            user_data["registration_date"] = "n/a"    

        try:
            user_data["last_visit"] = str(datetime.strptime(user.last_login.strftime('%d/%m/%Y'), '%d/%m/%Y'))
        except:
            try:
                user_data["last_visit"] = str(datetime.strptime(user.date_joined.strftime('%d/%m/%Y'), '%d/%m/%Y'))
            except:
                # should not occured
                user_data["last_visit"] = "n/a"
        
        try:
            user_data["country"] = json.loads(user.profile.custom_field)['country'].capitalize()
        except:
            user_data["country"] = "n/a"

        try:
            user_data["entity"] = json.loads(user.profile.custom_field)['entity'].capitalize()
        except:
            user_data["entity"] = "n/a"

        try:
            user_data["job"] = json.loads(user.profile.custom_field)['job'].capitalize()
        except:
            user_data["job"] = "n/a"

        try:
            user_data["langage"] = json.loads(user.profile.custom_field)['langage'].capitalize()
        except:
            user_data["langage"] = "n/a"

        user_data["id"] = user.id
        user_data["email_address"] = user.email

        # Time tracking
        log.info("tma_enrollment")
        log.info(tma_enrollment)
        log.info(tma_enrollment.global_time_tracking)
        try:
            seconds = tma_enrollment.global_time_tracking
            minute = seconds // 60
            user_data["time_tracking"] = int(minute)
        except:
            user_data["time_tracking"] = int(0)
            log.info("except 1")


        # Finished course date
        log.info("tma_enrollment.finished_course_date")
        log.info(tma_enrollment.finished_course_date)
        # try:
        #     seconds = tma_enrollment.finished_course_date
        #     log.info("try 3 ")

        #     minute = seconds // 60
        #     user_data["time_tracking"] = int(minute)
        # except:
        #     user_data["time_tracking"] = int(0)

        course_grade = CourseGradeFactory().create(user, course)

        locations_to_scores = (course_grade.chapter_grades[0]['sections'][0].locations_to_scores)
        
        # Access Section
        scorable_block_titles = OrderedDict()
        grading_context = grading_context_for_course(course_key)

        list_question = []


        for assignment_type_name, subsection_infos in grading_context['all_graded_subsections_by_type'].iteritems():
            for subsection_index, subsection_info in enumerate(subsection_infos, start=1):
                for scorable_block in subsection_info['scored_descendants']:
                    header_name = (
                        u"{assignment_type} {subsection_index}: "
                        u"{subsection_name} - {scorable_block_name}"
                    ).format(
                        scorable_block_name=scorable_block.display_name,
                        assignment_type=assignment_type_name,
                        subsection_index=subsection_index,
                        subsection_name=subsection_info['subsection_block'].display_name,
                    )
                    scorable_block_titles[scorable_block.location] = header_name
                    section_name = scorable_block.display_name
                    list_question.append(section_name)

        user_grade = check_best_grade(user, course, force_best_grade=True)
        log.info("user_grade")
        log.info(user_grade)


        user_state_client = DjangoXBlockUserStateClient()
        questions = []

        # Average calculation
        # score_sum = ""
        # d_o_c_sum = 0
        inverted_score = 0
        answered_total = 0
        iteration = 0


        for block_location,block_title in scorable_block_titles.items():
            question = {}
           
            # Add submit_time
            try:
                if user_grade.locations_to_scores.get(block_location):
                    history_entries = list(user_state_client.get_history(user.username, block_location))
                    value = history_entries[0].state.get('student_answers').values()[0]  # ----->  choice_2

                    corrected_value = value.split('_')[1]
                    corrected_value = int(corrected_value)
                    corrected_value += 1
                    answer = "choice "+ str(corrected_value)
                else:
                    answer=('not graded for student')
            except:
                answer='n.a.'

            # Add Timestamp

            try:
                if user_grade.locations_to_scores.get(block_location) and history_entries:
                    valueTS=history_entries[0].state.get('last_submission_time')

                    valueTS = str(valueTS)
                    dateList = valueTS.split("T")

                    date = dateList[0]
                    year = date.split('-')[0]
                    month = date.split('-')[1]
                    day = date.split('-')[2]

                    fulltime = dateList[1]
                    hour =fulltime.split(':')[0]
                    hour = int(hour) +2
                    minutes =fulltime.split(':')[1]
                    seconds =fulltime.split(':')[2]
                    seconds = seconds[0:1]

                    valueTS = str(day) + "/" +str(month) + "/" +str(year) + " "+str(hour)+"h:"+str(minutes) + "m:" +str(seconds)+ "s"

                else:
                    valueTS=('no time stamp')
            except:
                valueTS=('n.a.')

            valueTS = str(valueTS)
            # log.info(type(valueTS))
            # log.info((valueTS))
            # try:
            #     temps = datetime.strptime((valueTS),"%yyyy-%MM-%dd'T'%HH:%mm:%ss'Z'")
            #     log.info("temps")
            #     log.info(temps)
            # except:
            #     log.info("No timestamp")

            user_data["last_submit"] = valueTS

            grade = user_grade.chapter_grades[0]['sections'][0].scores[iteration]
            iteration += 1

            degree_table = {
                1 : {"graded":True, "inverted_matrix":20, "raw_score":1, "score":grade.earned, "choice":answer, "question":block_title, "submit_time":valueTS, "d_o_c": 0.975},
                0.975 : {"graded":True, "inverted_matrix":19, "raw_score":1, "score":grade.earned, "choice":answer, "question":block_title, "submit_time":valueTS, "d_o_c": 0.9},
                0.95 : {"graded":True, "inverted_matrix":18, "raw_score":1, "score":grade.earned, "choice":answer, "question":block_title, "submit_time":valueTS, "d_o_c": 0.775},
                0.925 : {"graded":True, "inverted_matrix":17, "raw_score":1, "score":grade.earned, "choice":answer, "question":block_title, "submit_time":valueTS, "d_o_c": 0.6},
                0.9 : {"graded":True, "inverted_matrix":16, "raw_score":1, "score":grade.earned, "choice":answer, "question":block_title, "submit_time":valueTS, "d_o_c": 0.375},
                0.825 : {"graded":True, "inverted_matrix":13, "raw_score":1, "score":grade.earned, "choice":answer, "question":block_title, "submit_time":valueTS, "d_o_c": 0.125},
                0 : {"graded":True, "inverted_matrix":4, "raw_score":0, "score":grade.earned, "choice":answer, "question":block_title, "submit_time":valueTS, "d_o_c": 0.975},
                0.35 : {"graded":True, "inverted_matrix":3, "raw_score":0, "score":grade.earned, "choice":answer, "question":block_title, "submit_time":valueTS, "d_o_c": 0.9},
                0.5 : {"graded":True, "inverted_matrix":2, "raw_score":0, "score":grade.earned, "choice":answer, "question":block_title, "submit_time":valueTS, "d_o_c": 0.775},
                0.55 : {"graded":True, "inverted_matrix":0, "raw_score":0, "score":grade.earned, "choice":answer, "question":block_title, "submit_time":valueTS, "d_o_c": 0.6},
                0.575 : {"graded":True, "inverted_matrix":-6, "raw_score":0, "score":grade.earned, "choice":answer, "question":block_title, "submit_time":valueTS, "d_o_c": 0.375},
                0.6 : {"graded":True, "inverted_matrix":-20, "raw_score":0, "score":grade.earned, "choice":answer, "question":block_title, "submit_time":valueTS, "d_o_c": 0.125}
            }

            if grade.earned in degree_table and answer !='n.a.': 
                question = degree_table[grade.earned]
                inverted_score += question["inverted_matrix"]
                answered_total += 1
                # score_sum += degree_table[grade.earned]["score"]
                # score_sum += str(degree_table[grade.earned]["score"])
                # score_sum += str(" + ")
                # d_o_c_sum += degree_table[grade.earned]["d_o_c"]
                
            else:
                question = {"graded": False, "inverted_matrix":0, "raw_score": 0, "score": 0,"choice":"0" ,"question":block_title,"submit_time": "n.a.", "d_o_c": 0}
                
            questions.append(question)


        # if len(score_sum) >= 2 :
        #     score_sum = score_sum[0: -2 ]

        
        if len(list_question) != 0 :

            user_data["average_score_raw"] = True
            user_data["average_score"] = True
            user_data["travail"] =  str( (answered_total) / len(list_question) )
            user_data["average_d_o_c"] = True
            user_data["centration"] = True
            user_data["average_inverted"] = round(int(inverted_score)/len(list_question), 2)

        else :
            user_data["average_score_raw"] = False
            user_data["average_score"] = False
            user_data["average_d_o_c"] = False
            user_data["centration"] = False
            user_data["average_inverted"] = False
        
        data = {"grades": questions, "general": user_data}

        all_users_data[str(user.id)]= data
        

headers = ["Identifiant","Email","Prénom", "Nom de famille","Pays","Entité","Job","Langage","Date d'inscription","Dernière connexion","Temps passé (hh-mm)","Date de soumission (hh-mm)", "Score global brut (en %)","Score global", "Degré de certitude moyen", "Centration" , "Matrice inversée"]


def Getletterfromindex( num):
    #produces a string from numbers so  1->A , 26->Z , 54->BB

    num2alphadict = dict(zip(range(1, 27), string.ascii_uppercase))
    outval = ""
    numloops = (num-1) //26
    if numloops > 0:
        outval = outval + Getletterfromindex(numloops)
    remainder = num % 26
    if remainder > 0:
        outval = outval + num2alphadict[remainder]
    else:
        outval = outval + "Z"
    return outval

# WRITE FILE
j = 2
i = 1
alphabet = []
for i, header in enumerate(headers):
    sheet.cell(j, i+1, header)

i=18
color = "00D0D0D0"

for question in list_question:
    # sheet.cell(0, i, "bonne réponse")
    # sheet.cell(0, i+1, "titre cours")
    sheet.cell(1, i, question).fill = PatternFill('solid', fgColor=color)
    sheet.cell(1, i+1).fill = PatternFill('solid', fgColor=color)
    sheet.cell(1, i+2).fill = PatternFill('solid', fgColor=color)
    sheet.cell(1, i+3).fill = PatternFill('solid', fgColor=color)
    sheet.cell(1, i+4).fill = PatternFill('solid', fgColor=color)
    sheet.cell(j, i, "Réponse étudiant").fill = PatternFill('solid', fgColor=color)
    sheet.cell(j, i+1, "Note brute").fill = PatternFill('solid', fgColor=color)
    sheet.cell(j, i+2, "Degré de certitude").fill = PatternFill('solid', fgColor=color)
    sheet.cell(j, i+3, "Score final").fill = PatternFill('solid', fgColor=color)
    sheet.cell(j, i+4, "Soumission").fill = PatternFill('solid', fgColor=color)

    i += 5 
    if color == "00D0D0D0" :
        color = "00FFFFFF"
    else:
        color = "00D0D0D0"



j = 2
for index, user in all_users_data.items():

    raw_score_sum = ""
    score_sum = ""
    d_o_c_sum = ""
    
    sheet.cell(j+1, 1, user["general"]["id"])
    sheet.cell(j+1, 2, user["general"]["email_address"])
    sheet.cell(j+1, 3, user["general"]["first_name"])
    sheet.cell(j+1, 4, user["general"]["last_name"])
    sheet.cell(j+1, 5, user["general"]["country"])
    sheet.cell(j+1, 6, user["general"]["entity"])
    sheet.cell(j+1, 7, user["general"]["job"])
    sheet.cell(j+1, 8, user["general"]["langage"])
    try:
        sheet.cell(j+1, 9, user["general"]["registration_date"])
        # cell.number_format = 'YYYY MMM DD'
    except:
        sheet.cell(j+1, 9, "n.a.")

    try:
        sheet.cell(j+1, 10, user["general"]["last_visit"])
        # sheet.cell(j+1, 9, user["general"]["last_visit"], style1)
    except:
        sheet.cell(j+1, 10, "n.a.")

    sheet.cell(j+1, 11, user["general"]["time_tracking"])
    try:
        sheet.cell(j+1, 12, user["general"]["last_submit"])
        # sheet.cell(j+1, 12, user["general"]["last_submit"], style1)
    except:
        sheet.cell(j+1, 12, 'no time stamp')

    i=18
    color = "00D0D0D0"
    for grade in user["grades"]:


        sheet.cell(j+1, i, grade["choice"]).fill = PatternFill('solid', fgColor=color)
        sheet.cell(j+1, i+1, grade["raw_score"]).fill = PatternFill('solid', fgColor=color)
        sheet.cell(j+1, i+2, grade["d_o_c"]).fill = PatternFill('solid', fgColor=color)
        sheet.cell(j+1, i+3, grade["score"]).fill = PatternFill('solid', fgColor=color)
        sheet.cell(j+1, i+4, grade["submit_time"]).fill = PatternFill('solid', fgColor=color)

        raw_score_sum += Getletterfromindex(i+1) + str(j+1)
        raw_score_sum += " + "
        d_o_c_sum += Getletterfromindex(i+2) + str(j+1)
        d_o_c_sum += " + "
        score_sum += Getletterfromindex(i+3) + str(j+1)
        score_sum += " + "

        i += 5

        if color == "00D0D0D0" :
            color = "00FFFFFF"
        else:
            color = "00D0D0D0"


    raw_score_sum = raw_score_sum[0:-2]
    d_o_c_sum = d_o_c_sum[0:-2]
    score_sum = score_sum[0:-2]

    if user["general"]["average_score_raw"] : 
        sheet.cell(j+1, 13, ("=SUM("+raw_score_sum+")*100 /" + str(len(user["grades"])) ))
    else:
        sheet.cell(j+1, 13, "n.a.")
    
    if user["general"]["average_score"] :
        sheet.cell(j+1, 14, ("=SUM("+score_sum +")*100 /" + str(len(user["grades"]))) )   
    else:
        sheet.cell(j+1, 14, "n.a.")   

    if user["general"]["average_d_o_c"] :
        sheet.cell(j+1, 15, ("=SUM("+d_o_c_sum + ")*100 /" + str(len(user["grades"]))))
    else: 
        sheet.cell(j+1, 15, "n.a.")

    if user["general"]["centration"] :
        # La centration est la différence entre le taux de bonne réponse (Nb de bonnes réponses/Nb de questions) et la certitude moyenne (degrée de certitude moyen. On prend comme référence pour une plage de certitude, la médiane. Par exemple le degrée de certitude 70%-85% deviens 77,5%
        sheet.cell(j+1, 16, ("=SUM("+(Getletterfromindex(13)+str(j+1))+"-"+(Getletterfromindex(15)+str(j+1))+")"))
    else :
        sheet.cell(j+1, 16, "n.a.")

    if user["general"]["average_inverted"] :
        sheet.cell(j+1, 17, user["general"]["average_inverted"])
    else :
        sheet.cell(j+1, 17, "n.a.")

    j += 1


# SEND MAILS
output = BytesIO()
wb.save(output)
_files_values = output.getvalue()

course_names = []
course_names_html = []
for course_id in course_ids: 
    course = get_course_by_id(CourseKey.from_string(course_id)) 
    course_names.append(course.display_name_with_default.encode('ascii', errors='xmlcharrefreplace'))
    course_names_html.append("<li>"+ course.display_name_with_default.encode('ascii', errors='xmlcharrefreplace')+"</li>")

course_names_html = ''.join(course_names_html)

html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en pièce jointe le rapport de note : "+ course_names_html +"<br/><br/></p></body></html>"

part2 = MIMEText(html.encode('utf-8'), 'html', 'utf-8')

for i in range(len(TO_EMAILS)):
   fromaddr = "ne-pas-repondre@themoocagency.com"
   toaddr = str(TO_EMAILS[i])
   msg = MIMEMultipart()
   msg['From'] = fromaddr
   msg['To'] = toaddr
   msg['Subject'] = "Engie - " + ' + '.join(course_names)
   attachment = _files_values
   part = MIMEBase('application', 'octet-stream')
   part.set_payload(attachment)
   encoders.encode_base64(part)
   part.add_header('Content-Disposition', "attachment; filename= %s" % os.path.basename(filename))
   msg.attach(part)
   server = smtplib.SMTP('mail3.themoocagency.com', 25)
   server.starttls()
   server.login('contact', 'waSwv6Eqer89')
   msg.attach(part2)
   text = msg.as_string()
   server.sendmail(fromaddr, toaddr, text)
   server.quit()
   log.info('Email sent to '+str(TO_EMAILS[i]))


 
# July 2021

# sudo -H -u edxapp /edx/bin/python.edxapp /edx/app/edxapp/edx-microsite/nuclear-training-program/utils/script-grade.py "cyril.adolf@weuplearning.com" "course-v1:nuclear-training-program+NTP26+Track17" 




