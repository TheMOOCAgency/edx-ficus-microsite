# -*- coding: utf-8 -*-
import sys
reload(sys)
sys.setdefaultencoding('utf8')

import os
import json
import csv
from xlwt import *
import time
import logging

from django.utils.translation import ugettext as _

from django.conf import settings

from django.http import Http404, HttpResponseServerError, HttpResponse
from util.json_request import JsonResponse
from student.models import User,CourseEnrollment,UserProfile
from openedx.core.djangoapps.site_configuration import helpers as configuration_helpers
from lms.djangoapps.tma_grade_tracking.models import dashboardStats
from tma_ensure_form.utils import ensure_form_factory
from lms.djangoapps.grades.new.course_grade import CourseGradeFactory
from opaque_keys.edx.locations import SlashSeparatedCourseKey
from opaque_keys.edx.keys import CourseKey
from courseware.courses import get_course_by_id
from openedx.core.djangoapps.course_groups.models import CohortMembership, CourseUserGroup
from openedx.core.djangoapps.course_groups.cohorts import get_cohort, is_course_cohorted
from tma_apps.models import TmaCourseEnrollment
import time
from collections import OrderedDict
from lms.djangoapps.grades.context import grading_context_for_course

from io import BytesIO

import smtplib
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText
from email.MIMEBase import MIMEBase
from email import encoders

from django.core.mail import EmailMessage
from microsite_configuration.models import Microsite
from django.conf import settings
from tma_apps.best_grade.helpers import check_best_grade
from courseware.user_state_client import DjangoXBlockUserStateClient
from openpyxl import Workbook as openpyxlWorkbook


log = logging.getLogger(__name__)

#OLD VERSION GRADES REPORT
class grade_reports():
    def __init__(self,request,course_id=None,microsite=None,filename=None,filepath=None,subscribe_report=False):
        self.request = request
        self.course_id = course_id
        self.microsite = microsite
        self.filename = filename
        self.filepath = filepath
        self.subscribe_report = subscribe_report

    def tma_get_scorable_blocks_titles(self, course_key):
        """
        Returns an dict that maps a scorable block's location id to its title.
        """
        scorable_block_titles = OrderedDict()
        grading_context = grading_context_for_course(course_key)
        for assignment_type_name, subsection_infos in grading_context['all_graded_subsections_by_type'].iteritems():
            for subsection_index, subsection_info in enumerate(subsection_infos, start=1):
                for scorable_block in subsection_info['scored_descendants']:
                    header_name = (
                        u"{assignment_type} {subsection_index}: "
                        u"{subsection_name} - {scorable_block_name}"
                    ).format(
                        scorable_block_name=scorable_block.display_name,
                        assignment_type=assignment_type_name,
                        subsection_index=subsection_index,
                        subsection_name=subsection_info['subsection_block'].display_name,
                    )
                    scorable_block_titles[scorable_block.location] = header_name
        return scorable_block_titles

    def get_time_tracking(self,enrollment):
        tma_enrollment,is_exist=TmaCourseEnrollment.objects.get_or_create(course_enrollment_edx=enrollment)
        seconds = tma_enrollment.global_time_tracking
        hour = seconds // 3600
        seconds %= 3600
        minute = seconds // 60
        global_time = str(hour)+"h"+str(minute)+"min"
        return global_time
    
    def getOldDatas(self):
        file = open(self.old_file, "rb")
        users_list=[]
        old_users_datas_list = csv.DictReader(file, delimiter=';')
        for user in old_users_datas_list:
            users_list.append(user)
        file.close()
        return users_list

    def task_generate_xls(self, old_file):
        #Get report infos
        self.old_file = old_file
        self.microsite = self.request.get('microsite')
        report_fields = self.request.get('form')
        register_fields = self.request.get('register_form')
        certificate_fields = self.request.get('certificate_form')
        select_custom_field_key=self.request.get('select_custom_field_key')
        select_custom_field_values=self.request.get('select_custom_field_values')
        split_by=self.request.get('split_by')
        do_not_send_email=self.request.get('do_not_send_email',False)
        old_users_datas_list = self.getOldDatas()
        
        log.info('Start task generate grade report for course {}'.format(self.course_id))

        if isinstance(self.course_id, list):
            course_key=CourseKey.from_string(self.course_id[0])
        else :
            course_key=CourseKey.from_string(self.course_id)
        course=get_course_by_id(course_key)
        microsite_information = Microsite.objects.get(key=self.microsite)

        form_factory = ensure_form_factory()
        form_factory.connect(db='ensure_form',collection='certificate_form')


        #Dict of labels
        form_labels={
            "last_connexion":_("Last login"),
            "inscription_date":_("Register date"),
            "user_id":_("User id"),
            "email":_("Email"),
            "grade_final":_("Final Grade"),
            "cohorte_names":_("Cohorte name"),
            "time_tracking":_("Time spent"),
            "certified":_("Attestation"),
            "username":_("Username"),
            "best_grade":_("Best Grade"),
            "best_grade_date":_("Best Grade Date"),
        }
        for field in register_fields :
            form_labels[field.get('name')]=field.get('label')
        for field in certificate_fields :
            form_labels[field.get('name')]=field.get('label')

        #Identify multiple cells fields
        multiple_cell_fields=["exercises_grade","grade_detailed","exercises_answers"]

        #Is report cohort specific?
        course_cohorted=is_course_cohorted(course_key)
        cohortes_targeted = []
        if course_cohorted :
            cohortes_targeted=[field.replace('cohort_selection_','') for field in report_fields if field.find('cohort_selection_')>-1]
            if cohortes_targeted and not 'cohorte_names' in report_fields:
                report_fields.append('cohorte_names')
        else :
            if 'cohorte_names' in report_fields:
                report_fields.remove('cohorte_names')

        #Get Graded block for exercises_grade details
        scorable_blocks_titles = self.tma_get_scorable_blocks_titles(course_key)

        #Create Workbook
        wb = openpyxlWorkbook()
        filename = '/home/edxtma/csv/{}_{}.xls'.format(time.strftime("%Y_%m_%d"),course.display_name_with_default)
        sheet =  wb.active
        sheet.title = "Grade Report"

        #Write information
        line=1
        course_enrollments=CourseEnrollment.objects.filter(course_id=course_key, is_active=1)
        for enrollment in course_enrollments :
            user= enrollment.user
            if not '@weuplearning.com' in user.email and not '@themoocagency.com' in user.email: 
                old_user_line = -1
                for idx, old_user in enumerate(old_users_datas_list):
                    if old_user["email"] == user.email:
                        old_user_line = idx
                log.info("-------------------------- treating user {} for grade report -------------------------".format(user.email))
                #do not include in reports if not active
                if not enrollment.is_active:
                    continue

                #Cohort Specific report - remove students not targetted
                if course_cohorted :
                    user_cohorte=get_cohort(user, course_key).name
                    if cohortes_targeted and not user_cohorte in cohortes_targeted :
                        continue

                #Custom field specific report - remove students not targetted
                #Get user custom field
                try:
                    custom_field = json.loads(UserProfile.objects.get(user=user).custom_field)
                except:
                    custom_field = {}
                if select_custom_field_key :
                    if select_custom_field_values and not custom_field.get(select_custom_field_key) in select_custom_field_values :
                        continue
                    else :
                        if not select_custom_field_key in report_fields:
                            report_fields.append(select_custom_field_key)

                #Gather user information
                custom_user_grade = 0
                user_grade = []
                grade_summary={}
                if not self.subscribe_report:
                    user_grade = check_best_grade(user, course, force_best_grade=("best_grade" in report_fields))
                    for section_grade in user_grade.grade_value['section_breakdown']:
                        grade_summary[section_grade['category']]=section_grade['percent']

                tma_enrollment=TmaCourseEnrollment.get_enrollment(course_id=self.course_id, user=user)
                user_certificate_info = {}
                try:
                    form_factory.microsite = self.microsite
                    form_factory.user_id = user.id
                    user_certificate_info = form_factory.getForm(user_id=True,microsite=True).get('form')
                except:
                    pass

                user_state_client = DjangoXBlockUserStateClient()

                cell=1
                for field in report_fields :
                    if field in multiple_cell_fields:
                        if field=="grade_detailed":
                            for section in sorted(grade_summary):
                                grade_value = 0
                                grade_value = grade_summary[section]
                                custom_user_grade +=grade_value
                                section_grade = str(int(round(grade_value * 100)))+'%'
                                sheet.cell(row=line+1, column=cell, value=section_grade)
                                if line ==1 :
                                    sheet.cell(row=1, column=cell, value="Travail - "+section)
                                cell+=1
                        elif field=="exercises_grade":
                            for block_location,block_title  in scorable_blocks_titles.items():
                                if block_title == "Société Civile 1: Dispositif de prévention sur la lutte anti-blanchiment - Quiz" or block_title == "Société Civile 5: Évaluation Thématique 4 - Quiz":
                                    try:
                                        if user_grade.locations_to_scores.get(block_location):
                                            block_detail = user_grade.locations_to_scores.get(block_location)
                                            data_in_old_file = False
                                            if old_user_line > -1 and str(block_title) in old_users_datas_list[old_user_line]:
                                                data_in_old_file = True
                                            if (data_in_old_file and old_users_datas_list[old_user_line][str(block_title)] != "0") or (data_in_old_file and old_users_datas_list[old_user_line][str(block_title)] and block_detail.attempted and round(float(old_users_datas_list[old_user_line][str(block_title)])) > round(float(block_detail.earned)/block_detail.possible, 2)):
                                                value = old_users_datas_list[old_user_line][str(block_title)]
                                                if (data_in_old_file and old_users_datas_list[old_user_line][str(block_title)] and block_detail.attempted and round(float(old_users_datas_list[old_user_line][str(block_title)])) > round(float(block_detail.earned)/block_detail.possible, 2)):
                                                    log.info("***************** old grade's value for "+block_title+" is better ("+str(value)+")+than current ("+str(round(float(block_detail.earned)/block_detail.possible, 2))+") course for user : "+str(user.email))
                                            elif block_detail.attempted:
                                                value=round(float(block_detail.earned)/block_detail.possible, 2)
                                            else:
                                                value=_('n.a.')
                                        else :
                                            value=('not graded for student')
                                    except:
                                        value=_('inv.')
                                    sheet.cell(row=line+1, column=cell, value=value)
                                    if line==1 :
                                        sheet.cell(row=1, column=cell, value="Grade - "+block_title)
                                    cell+=1
                                    if "exercises_answers" in report_fields:
                                        #Answer
                                        try:
                                            if user_grade.locations_to_scores.get(block_location):
                                                history_entries = list(user_state_client.get_history(user.username, block_location))
                                                value = history_entries[0].state.get('student_answers').values()[0]
                                                if isinstance(value, basestring) and "choice_" in value:
                                                    value=self.addOneToChoice(value)
                                                elif isinstance(value, basestring) and not "choice_" in value:
                                                    value = "diff type"
                                                elif isinstance(value, list):
                                                    value=[self.addOneToChoice(choice) for choice in value]
                                                    value=", ".join(value)
                                            else:
                                                value=('not graded for student')
                                        except:
                                            value=_('inv.')
                                        sheet.cell(row=line+1, column=cell, value=value)
                                        if line==1 :
                                            sheet.cell(row=1, column=cell, value="Answers - "+scorable_blocks_titles[block_location])
                                        cell+=1
                                        #Submission Stamp
                                        try:
                                            if user_grade.locations_to_scores.get(block_location) and history_entries:
                                                value=history_entries[0].state.get('last_submission_time')
                                            else:
                                                value=('no time stamp')
                                        except:
                                            value=_('inv.')
                                        sheet.cell(row=line+1, column=cell, value=value)
                                        if line==1 :
                                            sheet.cell(row=1, column=cell, value="Last submission - "+scorable_blocks_titles[block_location])
                                        cell+=1
                                        history_entries=None

                        elif field=="exercises_answers" and not "exercises_grade" in report_fields:
                            for block_location,block_title  in scorable_blocks_titles.items():
                                #Answer
                                try:
                                    if user_grade.locations_to_scores.get(block_location):
                                        history_entries = list(user_state_client.get_history(user.username, block_location))
                                        value = history_entries[0].state.get('student_answers').values()[0]
                                        if isinstance(value, basestring) and "choice_" in value:
                                            value=self.addOneToChoice(value)
                                        elif isinstance(value, basestring) and not "choice_" in value:
                                            value = "diff type"
                                        elif isinstance(value, list):
                                            value=[self.addOneToChoice(choice) for choice in value]
                                            value=", ".join(value)
                                    else:
                                        value=('not graded for student')
                                except:
                                    value=_('inv.')
                                sheet.cell(row=line+1, column=cell, value=value)
                                if line==1 :
                                    sheet.cell(row=1, column=cell, value="Answers - "+scorable_blocks_titles[block_location])
                                cell+=1
                                #Submission Stamp
                                try:
                                    if user_grade.locations_to_scores.get(block_location) and history_entries:
                                        value=history_entries[0].state.get('last_submission_time')
                                    else:
                                        value=('no time stamp')
                                except:
                                    value=_('inv.')
                                sheet.cell(row=line+1, column=cell, value=value)
                                if line==1 :
                                    sheet.cell(row=1, column=cell, value="Last submission - "+scorable_blocks_titles[block_location])
                                cell+=1
                                history_entries=None
                    else :
                        value=''
                        if field=="user_id":
                            value=user.id
                        elif field=="email":
                            value=user.email
                        elif field=="first_name":
                            try :
                                if user.first_name:
                                    value=user.first_name
                                elif custom_field :
                                    value=custom_field.get('first_name', 'unkowna')
                                else :
                                    value='unknown'
                            except :
                                value='unknown'
                        elif field=="last_name":
                            try :
                                if user.last_name:
                                    value=user.last_name
                                elif custom_field:
                                    value=custom_field.get('last_name', 'unkowna')
                            except :
                                value='unknown'
                        elif field=="last_connexion":
                            try :
                                value=user.last_login.strftime('%d-%m-%y')
                            except:
                                value=''
                        elif field=="inscription_date":
                            try :
                                value=user.date_joined.strftime('%d-%m-%y')
                            except:
                                value=''
                        elif field=="cohorte_names":
                            try:
                                value=user_cohorte
                            except:
                                value=''
                        elif field=="time_tracking":
                            value=self.get_time_tracking(enrollment)
                        elif field=="best_grade":
                            try:
                                value = str(int(round(user_grade.percent_tma*100)))+"%"
                            except:
                                value='n.a'
                        elif field=="best_grade_date":
                            try:
                                value=tma_enrollment.best_grade_date.strftime('%d-%m-%y')
                            except:
                                value='n.a'
                        elif field=="certified":
                            value = _("Yes")
                        elif field=="grade_final":
                            #value = str(int(round(user_grade.percent * 100)))+'%'
                            value = str(int(round((custom_user_grade / 4) * 100)))+'%'
                        elif field=="username":
                            value=user.username
                        elif field in user_certificate_info.keys():
                            value=user_certificate_info.get(field)
                        else :
                            value=custom_field.get(field,'')
                            if value == '' and old_user_line > -1 and field in old_users_datas_list[old_user_line]:  
                                value = old_users_datas_list[old_user_line][field]
                        #Write header and write value
                        if field in form_labels.keys():
                            sheet.cell(row=line+1, column=cell, value=value)
                            if line==1:
                                sheet.cell(row=1, column=cell, value=form_labels.get(field))
                            cell+=1
                line+=1
                if old_user_line > -1 :
                    del old_users_datas_list[old_user_line]
            else :
                log.info("-------------------------- skip admin user {} for grade report -------------------------".format(user.email))
        for old_user in old_users_datas_list :
            log.info("-------------------------- treating OLD user {} for grade report -------------------------".format(old_user["email"]))
            cell=1
            sum_grade=0
            for field in report_fields :
                value=''
                if field in multiple_cell_fields:
                    if field=="grade_detailed":
                        for section in sorted(grade_summary):
                            section_grade = 'n/a'
                            sheet.cell(row=line+1, column=cell, value=section_grade)
                            cell+=1
                else :
                    if field=="email":
                        value=old_user[field]
                    elif field=="username":
                        value=old_user[field]
                    elif field=="first_name":
                        value=old_user["firstname"]
                    elif field=="last_name":
                        value=old_user["lastname"]
                    elif field=="registration_number":
                        value=old_user["matricule"]
                    elif field=="manager":
                        value=old_user[field]
                    elif field=="site_name":
                        value=old_user[field]
                    elif field=="job_title":
                        value=old_user[field]
                    elif field=="last_connexion":
                        value=old_user["derniere connexion"]
                    elif field=="inscription_date":
                        value=old_user["inscrit le"]
                    elif field=="best_grade_date":
                        value=old_user[field]
                    elif field=="grade_final":
                        value = str(int(round((sum_grade/4) * 100)))+'%'
                    elif field=="certified":
                        value="Oui"
                    else :
                        value="n.a"
                    sheet.cell(row=line+1, column=cell, value=value)
                    cell+=1
            line+=1


        #Save the file
        output = BytesIO()
        wb.save(output)
        _files_values = output.getvalue()
        log.warning("file saved")

        #If no email is to be sent, then just send the response to the caller as a BytesIO object
        if do_not_send_email:
            response = {
                'path':self.filename,
                'send_to':receivers,
                'xls_file':_files_values
            }
            return response

        #Send the email to receivers
        receivers = self.request.get('send_to')

        if cohortes_targeted and len(cohortes_targeted)>1:
            html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en PJ le rapport de donnees du MOOC {} pour les cohortes {}<br/><br/>Si vous disposez d'accès suffisants vous pouvea accéder au dashboard du cours: https://{}/tma/{}/dashboard <br><br> et au studio du cours : https://{}/course/{}    <br/><br/>Bonne reception<br>The MOOC Agency<br></p></body></html>".format(course.display_name, ' , '.join(cohortes_targeted), microsite_information.values['SITE_NAME'], course.id, settings.CMS_BASE, course.id)
        elif cohortes_targeted and len(cohortes_targeted)==1:
            html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en PJ le rapport de donnees du MOOC {} pour la cohorte {}<br/><br/>Si vous disposez d'accès suffisants vous pouvea accéder au dashboard du cours: https://{}/tma/{}/dashboard <br><br> et au studio du cours : https://{}/course/{}    <br/><br/>Bonne reception<br>The MOOC Agency<br></p></body></html>".format(course.display_name, ' '.join(cohortes_targeted), microsite_information.values['SITE_NAME'], course.id, settings.CMS_BASE, course.id)
        else :
            html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en PJ le rapport de donnees du MOOC {}<br/><br/>Si vous disposez d'accès suffisants vous pouvea accéder au dashboard du cours: https://{}/tma/{}/dashboard <br><br> et au studio du cours : https://{}/course/{}    <br/><br/>Bonne reception<br>The MOOC Agency<br></p></body></html>".format(course.display_name, microsite_information.values['SITE_NAME'], course.id, settings.CMS_BASE, course.id)
        part2 = MIMEText(html.encode('utf-8'), 'html', 'utf-8')

        for receiver in receivers :
            fromaddr = "ne-pas-repondre@themoocagency.com"
            toaddr = str(receiver)
            msg = MIMEMultipart()
            msg['From'] = fromaddr
            msg['To'] = toaddr

            subject = "{} - Rapport de donnees".format(course.display_name)
            if cohortes_targeted:
                subject += ' - Filtre Cohortes :'+' '.join(cohortes_targeted)
            if select_custom_field_key and select_custom_field_values:
                log.info(select_custom_field_values)
                log.info(' '.join(select_custom_field_values))
                subject += (' - Filtre : '+' '.join(select_custom_field_values))

            msg['Subject'] = subject

            attachment = _files_values
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', "attachment; filename= %s" % os.path.basename(filename))
            msg.attach(part)
            server = smtplib.SMTP('mail3.themoocagency.com', 25)
            server.starttls()
            server.login('contact', 'waSwv6Eqer89')
            msg.attach(part2)
            text = msg.as_string()
            server.sendmail(fromaddr, toaddr, text)
            server.quit()
            log.warning("file sent to {}".format(receiver))

        response = {
            'path':self.filename,
            'send_to':receivers
        }

        return response

    def addOneToChoice(self, choice):
        choice=choice.split("_")
        choice[1]=str(int(choice[1])+1)
        return "_".join(choice)

    def download_xls(self):
        self.filepath = '/edx/var/edxapp/grades/{}'.format(self.filename)
        _file = open(self.filepath,'r')
        _content = _file.read()
        response = HttpResponse(_content, content_type="application/vnd.ms-excel")
        response['Content-Disposition'] = "attachment; filename="+self.filename
        return response
