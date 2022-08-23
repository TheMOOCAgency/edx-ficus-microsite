# -*- coding: utf-8 -*-
#!/usr/bin/env python
import sys

reload(sys)
sys.setdefaultencoding('utf8')

import os
import importlib
import csv
import json


os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lms.envs.aws")
os.environ.setdefault("lms.envs.aws,SERVICE_VARIANT", "lms")
os.environ.setdefault("PATH", "/edx/app/edxapp/venvs/edxapp/bin:/edx/app/edxapp/edx-platform/bin:/edx/app/edxapp/.rbenv/bin:/edx/app/edxapp/.rbenv/shims:/edx/app/edxapp/.gem/bin:/edx/app/edxapp/edx-platform/node_modules/.bin:/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin")

os.environ.setdefault("SERVICE_VARIANT", "lms")
os.chdir("/edx/app/edxapp/edx-platform")

startup = importlib.import_module("lms.startup")
startup.run()

from django.core.management import execute_from_command_line

from django.contrib.auth.models import User
from courseware.courses import get_course_by_id
from microsite_configuration.models import Microsite
from itertools import repeat
from tma_apps.models import TmaCourseEnrollment
from tma_apps.skill_grade.helpers import SkillGrades

from openpyxl import Workbook


import logging
log = logging.getLogger()


# Define variables
microsite_name = sys.argv[1]
users_file_name = sys.argv[2]

structure_fields=['struct_org1','struct_org2','struct_org3','struct_org4','struct_org5','struct_org6','struct_org7','struct_org8','struct_org9']


# usernames_to_exclude = ['', 'testtmaviewableall', 'testtmagermanmanager']

# users_right_file = json.load(f)
f = open('/edx/var/edxapp/secret/microsite/'+microsite_name+'/'+users_file_name)
usernames_list = csv.DictReader(f, delimiter=';')

g = open('/edx/var/edxapp/secret/microsite/'+microsite_name+'/localisation_dict.json',)
localisation_dict = json.load(g)

h = open('/edx/var/edxapp/secret/microsite/'+microsite_name+'/position_dict.json',)
position_dict = json.load(h)

# TRY TO USE GRADES.JSON
l = open('/edx/var/edxapp/secret/microsite/'+microsite_name+'/grades.json',)
grades = json.load(l)
print("grades.json is loaded")



def buildGradesFile(user_from_list):

    list_of_global_grade=[]
    list_of_skill_grade=[]
    has_finished = []
    has_started = []
    found_at_previous_level = True

    for level in range(4):
        if not found_at_previous_level :
            break

        found_at_previous_level = False
        for user in grades[level]['userData']:

            if user['fields']['id'] == user_from_list['Uid']:
                # user IN GRADES.JSON
                print("*********** grade founded ************")
                print(user['grades'])
                list_of_global_grade.append(user['grades']['global'])
                list_of_skill_grade.append(user['grades']['skills'])
                found_at_previous_level = True
                has_finished.append(user['fields']['has_finished'])
                has_started.append(user['fields']['has_started'])
                break

        # user NOT IN GRADES.JSON. FOR THE CURRENT level
        if level == 0 and list_of_skill_grade == []:
            print('not enrolled')
            break


    userResults={
        "fields":{
            "id":user_from_list['Uid'],
            "first_name":user_from_list['first_name'].lower(),
            "last_name":user_from_list['last_name'].lower(),
            "localisation":[user_from_list['country'],user_from_list['location']],
            "position":[user_from_list['job_family'],user_from_list['profession']],
            "structure":[user_from_list[structure] for structure in structure_fields],
            "has_finished":has_finished,
            "has_started":has_started,
            "grade_global":list_of_global_grade if list_of_global_grade != [] else False,
            "skills" : list_of_skill_grade if list_of_skill_grade != [] else False
        },
    }

    return userResults



# The admin list need to be updated in the react skill_dashboard.html
headers_admin = ['id', 'Prénom', 'Nom', 'Commencé 1', 'Terminé 1', 'Note globale level 1', 'Stratégie', 'Clients', 'Opération', 'Collaborateurs', 'Technologies', 'Commencé 2', 'Terminé 2', 'Note globale level 2', 'Stratégie', 'Clients', 'Opération', 'Collaborateurs', 'Technologies', 'Commencé 3', 'Terminé 3', 'Note globale level 3', 'Stratégie', 'Clients', 'Opération', 'Collaborateurs', 'Technologies', 'Commencé 4', 'Terminé 4', 'Note globale level 4', 'Stratégie', 'Clients', 'Opération', 'Collaborateurs', 'Technologies', 'localisation_0', 'localisation_1', 'position_0', 'position_1', 'structure_0', 'structure_1', 'structure_2', 'structure_3', 'structure_4', 'structure_5', 'structure_6', 'structure_7', 'structure_8']



wb = Workbook()
sheet = wb.active
sheet.title = 'Rapport_admin'

for i, header in enumerate(headers_admin):
    column = i+1
    sheet.cell(row=1, column=column, value=header)

j = 2

for user_from_list in usernames_list:

    user = buildGradesFile(user_from_list)
    sheet.cell(row=j, column=1, value=user['fields']["id"])
    sheet.cell(row=j, column=2, value=user['fields']["first_name"])
    sheet.cell(row=j, column=3, value=user['fields']["last_name"])
    sheet.cell(row=j, column=36, value=localisation_dict[user['fields']["localisation"][0]]) if user['fields']["localisation"][0] in localisation_dict else sheet.cell(row=j, column=36, value=user['fields']["localisation"][0])
    sheet.cell(row=j, column=37, value=localisation_dict[user['fields']["localisation"][1]]) if  user['fields']["localisation"][1] in localisation_dict else sheet.cell(row=j, column=37, value=user['fields']["localisation"][1])
    sheet.cell(row=j, column=38, value=position_dict[user['fields']["position"][0]]) if user['fields']["position"][0] in position_dict else sheet.cell(row=j, column=38, value=user['fields']["position"][0])
    sheet.cell(row=j, column=39, value=position_dict[user['fields']["position"][1]]) if user['fields']["position"][1] in position_dict else sheet.cell(row=j, column=39, value=user['fields']["position"][1])

    i = 40
    for struct in user['fields']["structure"]:
        sheet.cell(row=j, column=i, value=struct)
        i += 1

    # LOOP OVER GRADE AND SKILL FOR MULTI COURSE 
    # LOOP OVER GRADE AND SKILL FOR MULTI COURSE 
    k = 4
    for level in range(4):

        if user['fields']["grade_global"] :
            if len(user['fields']["grade_global"]) >= level +1 :

                sheet.cell(row=j, column=k, value='Vrai') if user['fields']["has_started"][level] else sheet.cell(row=j, column=k, value='Faux')
                sheet.cell(row=j, column=k+1, value='Vrai') if user['fields']["has_finished"][level] else sheet.cell(row=j, column=k+1, value='Faux')
                sheet.cell(row=j, column=k+2, value=user['fields']["grade_global"][level]) 
                # sheet.cell(row=j, column=k, user['fields']["grade_global"][level]) if user['fields']["grade_global"] else  sheet.cell(row=j, column=k, '-')

                # sheet.cell(row=j, column=k+1, user['fields']["skills"][level][1]['grade']) if user['fields']["skills"] else  sheet.cell(row=j, column=k+1, '-')
                sheet.cell(row=j, column=k+3, value=user['fields']["skills"][level][1]['grade'])
                sheet.cell(row=j, column=k+4, value=user['fields']["skills"][level][3]['grade'])
                sheet.cell(row=j, column=k+5, value=user['fields']["skills"][level][0]['grade'])
                sheet.cell(row=j, column=k+6, value=user['fields']["skills"][level][2]['grade'])
                sheet.cell(row=j, column=k+7, value=user['fields']["skills"][level][4]['grade'])

            else:
                sheet.cell(row=j, column=k, value='-')
                sheet.cell(row=j, column=k+1, value='-')
                sheet.cell(row=j, column=k+2, value='-')
                sheet.cell(row=j, column=k+3, value='-')
                sheet.cell(row=j, column=k+4, value='-')
                sheet.cell(row=j, column=k+5, value='-')
                sheet.cell(row=j, column=k+6, value='-')
                sheet.cell(row=j, column=k+7, value='-')
        else:
            sheet.cell(row=j, column=k, value='-')
            sheet.cell(row=j, column=k+1, value='-')
            sheet.cell(row=j, column=k+2, value='-')
            sheet.cell(row=j, column=k+3, value='-')
            sheet.cell(row=j, column=k+4, value='-')
            sheet.cell(row=j, column=k+5, value='-')
            sheet.cell(row=j, column=k+6, value='-')
            sheet.cell(row=j, column=k+7, value='-')

        k += 8

    print('end Level')
    j += 1

filename = "scope_specific_report_admin.xlsx"
filepath = '/edx/var/edxapp/media/microsite/{}/dtl_reports/{}'.format(microsite_name,filename)

wb.save(filepath)


log.info("END OF THE ADMIN SPECIFIC REPORT SCRIPT")



# List of command to execute:
# sudo -H -u edxapp /edx/bin/python.edxapp /edx/var/edxapp/secret/microsite/psa-netexplo/dtl_reports_script/admin_report_script_from_grades.py "psa-netexplo" "culture_digitale_export.csv"
 
# sudo -H -u edxapp /edx/bin/python.edxapp /edx/var/edxapp/secret/microsite/psa-netexplo/dtl_reports_script/admin_report_script_from_grades.py "psa-netexplo" "culture_digitale_export_small.csv"
