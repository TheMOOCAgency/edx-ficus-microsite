# -*- coding: utf-8 -*-
#!/usr/bin/env python
import sys
reload(sys)
sys.setdefaultencoding('utf8')

import os
import importlib
import time
import json

# Workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

from io import BytesIO

import smtplib
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText
from email.MIMEBase import MIMEBase
from email import encoders


os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lms.envs.aws")
os.environ.setdefault("lms.envs.aws,SERVICE_VARIANT", "lms")
os.environ.setdefault("PATH", "/edx/app/edxapp/venvs/edxapp/bin:/edx/app/edxapp/edx-platform/bin:/edx/app/edxapp/.rbenv/bin:/edx/app/edxapp/.rbenv/shims:/edx/app/edxapp/.gem/bin:/edx/app/edxapp/edx-platform/node_modules/.bin:/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin")
os.environ.setdefault("SERVICE_VARIANT", "lms")
os.chdir("/edx/app/edxapp/edx-platform")

startup = importlib.import_module("lms.startup")
startup.run()

from opaque_keys.edx.keys import CourseKey
from lms.djangoapps.grades.new.course_grade import CourseGradeFactory
from tma_apps.models import TmaCourseEnrollment
from student.models import *
from courseware.courses import get_course_by_id
from microsite_configuration.models import (
    MicrositeOrganizationMapping,
    Microsite
)
from openedx.core.djangoapps.course_groups.cohorts import get_cohort

import logging
log = logging.getLogger()


# ARGS
string_emails = sys.argv[1]
TO_EMAILS = string_emails.split(';')

course_id_list = sys.argv[2].split(';')


course_id = course_id_list[0]
fragmented_course_id = course_id.split('+')
lang = fragmented_course_id[1]


#get microsite
course_key = CourseKey.from_string(course_id_list[0])
course = get_course_by_id(course_key)
org = course.org
query = "SELECT a.id,a.organization,b.key FROM microsite_configuration_micrositeorganizationmapping a,microsite_configuration_microsite b WHERE a.microsite_id = b.id"
microsite_list = MicrositeOrganizationMapping.objects.raw(query)
microsite_name = None
for row in microsite_list:
    if row.organization == org:
        microsite_name = row.key


timestr = time.strftime("%Y_%m_%d")



course_enrollment = []

#get course enrolls

#grades reports summary
# user_repports_summary=course_enrollment[0]
# user_summary=user_repports_summary.user

wb = Workbook()
filename = '/home/edxtma/csv/{}_LVMH_report_{}.xlsx'.format(timestr, lang)
sheet = wb.active


def convert_second_to_time(seconds) :
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    return str(hours) + "h" + str(minutes) + "min"

#headers
HEADERS = ["E-mail","Prénom","Nom","Date d'inscription","Dernière connexion","Temps passé","Note finale","Status","Domaine","Cohorte"]
for i, header in enumerate(HEADERS):
    sheet.cell(1, i+1, header)
    sheet.cell(1, i+1).fill = PatternFill("solid", fgColor="191970")
    sheet.cell(1, i+1).font = Font(b=False, color="FFFFFF")



#NOW LOOK AT GRADES
j=1


for course_id in course_id_list :

    course_key = CourseKey.from_string(course_id)
    course = get_course_by_id(course_key)

    course_enrollment = CourseEnrollment.objects.filter(course_id=course_key)


    for i in range(len(course_enrollment)):

        enrollment = course_enrollment[i]

        user = course_enrollment[i].user
        email = user.email

        tma_enrollment, is_exist = TmaCourseEnrollment.objects.get_or_create(course_enrollment_edx=enrollment)
        time_tracking = convert_second_to_time(tma_enrollment.global_time_tracking)

        if (email.find('@weuplearning') != -1 or email.find('@yopmail') != -1 or email.find('@themoocagency.com') != -1) :
            continue

        last_login = user.last_login
        if last_login :
            j=j+1

            username=user.username
            user_profile = json.loads(UserProfile.objects.get(user=user).custom_field)

            try:
                first_name = user.first_name
            except:
                try: 
                    first_name = user_profile['first_name']
                except:
                    first_name = ""

            try:
                last_name = user.last_name
            except:
                try: 
                    last_name = user_profile['last_name']
                except:
                    last_name = ""

            try:
                user_cohorte = get_cohort(user, course_key)
                cohort_name = user_cohorte.name
            except: 
                cohort_name = 'Groupe par défaut'


            date_inscription = user.date_joined.strftime('%d/%m/%Y')
            last_login=user.last_login.strftime('%d/%m/%Y')
            client = email.split('@')[1]

            primary_rows = [
                email,
                first_name,
                last_name,
                date_inscription, 
                last_login,
                time_tracking
            ]

            last_rows = [
                client,
                cohort_name
            ]

            # Start to cell the line
            l=1

            for prim_row in primary_rows:
                sheet.cell(j, l, prim_row)
                l=l+1


            # -> quiz ptincipal + quiz alternatif mais une seule note, on regarde donc percent qui est la note globale
            course_grade = CourseGradeFactory().create(user, course)

            # Bug with percent over 100%
            percent = str(course_grade.percent * 100) + '%'
            if course_grade.percent > 1 :
                percent = str(course_grade.percent * 10) + '%'

            sheet.cell(j, l, percent.replace('.',','))
            l=l+1

            if int(percent.split('.')[0]) >= 75 : 
                sheet.cell(j, l, 'COMPLETED')
                sheet.cell(j, l).fill = PatternFill("solid", fgColor="008000")
                sheet.cell(j, l).font = Font(b=False, color="FFFFFF")
            else:
                sheet.cell(j, l, 'NOT COMPLETED')
                sheet.cell(j, l).fill = PatternFill("solid", fgColor="FF0000")
                sheet.cell(j, l).font = Font(b=False, color="FFFFFF")

            l=l+1

            for last_row in last_rows:
                sheet.cell(j, l, last_row)
                l=l+1


output = BytesIO()
wb.save(output)
_files_values = output.getvalue()

html = "<html><head></head><body><p>Bonjour,<br/><br/>Vous trouverez en PJ le rapport de données du MOOC {} ({})<br/><br/>Bonne réception<br>L'équipe WeUp Learning<br></p></body></html>".format(course.display_name, lang)
part2 = MIMEText(html.encode('utf-8'), 'html', 'utf-8')

for i in range(len(TO_EMAILS)):
   fromaddr = "ne-pas-repondre@themoocagency.com"
   toaddr = str(TO_EMAILS[i])
   msg = MIMEMultipart()
   msg['From'] = fromaddr
   msg['To'] = toaddr
   msg['Subject'] = "Rapport de donnees"
   attachment = _files_values
   part = MIMEBase('application', 'octet-stream')
   part.set_payload(attachment)
   encoders.encode_base64(part)
   part.add_header('Content-Disposition', "attachment; filename= %s" % os.path.basename(filename))
   msg.attach(part)
   server = smtplib.SMTP('mail3.themoocagency.com', 25)
   server.starttls()
   server.login('contact', 'waSwv6Eqer89')
   msg.attach(part2)
   text = msg.as_string()
   server.sendmail(fromaddr, toaddr, text)
   server.quit()
   print('mail send to '+str(TO_EMAILS[i]))

# TEST 
# sudo -H -u edxapp /edx/bin/python.edxapp /edx/app/edxapp/edx-microsite/lvmh-elearninginternalcontrol/utils/report_script.py "cyril.adolf@weuplearning.com" "course-v1:lvmh-elearninginternalcontrol+FR+2021;course-v1:lvmh-elearninginternalcontrol+FR+2022"

# CRONTAB
# 0 7 * * MON sudo -H -u edxapp /edx/bin/python.edxapp /edx/app/edxapp/edx-microsite/lvmh-elearninginternalcontrol/utils/report_script.py "emma.bekaert@lvmh.com;maico.naidedasilva@lvmh.com;manal.touati@weuplearning.com" course-v1:lvmh-elearninginternalcontrol+EN+2022 
# 0 7 * * MON sudo -H -u edxapp /edx/bin/python.edxapp /edx/app/edxapp/edx-microsite/lvmh-elearninginternalcontrol/utils/report_script.py "emma.bekaert@lvmh.com;maico.naidedasilva@lvmh.com;manal.touati@weuplearning.com" course-v1:lvmh-elearninginternalcontrol+CN+2022 
# 0 7 * * MON sudo -H -u edxapp /edx/bin/python.edxapp /edx/app/edxapp/edx-microsite/lvmh-elearninginternalcontrol/utils/report_script.py "emma.bekaert@lvmh.com;maico.naidedasilva@lvmh.com;manal.touati@weuplearning.com" course-v1:lvmh-elearninginternalcontrol+FR+2022 

